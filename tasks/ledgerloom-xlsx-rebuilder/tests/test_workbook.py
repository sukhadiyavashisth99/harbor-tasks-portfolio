import os
from decimal import Decimal
from openpyxl import load_workbook

OUT_PATH = "/output/final_report.xlsx"

def _load():
    assert os.path.exists(OUT_PATH), "Expected /output/final_report.xlsx to exist"
    return load_workbook(OUT_PATH, data_only=True)

def _sheet(wb, name):
    assert name in wb.sheetnames, f"Missing required sheet: {name}"
    return wb[name]

def test_output_file_exists():
    assert os.path.exists(OUT_PATH)

def test_transactions_sheet_exists():
    wb = _load()
    _sheet(wb, "Transactions")

def test_summary_sheet_exists():
    wb = _load()
    _sheet(wb, "Summary")

def test_transactions_headers_exact_order():
    wb = _load()
    ws = _sheet(wb, "Transactions")
    headers = [ws.cell(row=1, column=i).value for i in range(1, 6)]
    assert headers == ["txn_id", "date", "description", "category", "amount"]

def test_date_normalization_iso_format():
    wb = _load()
    ws = _sheet(wb, "Transactions")
    for r in [2, 3, 4, 6]:
        v = ws.cell(row=r, column=2).value
        assert isinstance(v, str)
        assert len(v) == 10 and v[4] == "-" and v[7] == "-", f"Bad ISO date: {v}"

def test_category_mapping_applied():
    wb = _load()
    ws = _sheet(wb, "Transactions")
    cats = set()
    for r in range(2, ws.max_row + 1):
        cats.add(ws.cell(row=r, column=4).value)
    assert "Groceries" in cats
    assert "Groc." not in cats

def test_summary_totals_match_expected():
    wb = _load()
    ws = _sheet(wb, "Summary")
    totals = {}
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        if cat is None:
            continue
        totals[str(cat)] = Decimal(str(val)).quantize(Decimal("0.01"))

    expected = {
        "Food": Decimal("0.00"),
        "Groceries": Decimal("-42.37"),
        "Health": Decimal("-63.98"),
        "Income": Decimal("2850.00"),
        "Other": Decimal("-12.49"),
        "Rent": Decimal("-1200.00"),
        "Transport": Decimal("-53.71"),
        "Utilities": Decimal("-96.13"),
    }

    for k, v in expected.items():
        assert k in totals
        assert totals[k] == v

def test_grand_total_row_present_and_correct():
    wb = _load()
    ws = _sheet(wb, "Summary")
    last_cat = ws.cell(row=ws.max_row, column=1).value
    last_val = ws.cell(row=ws.max_row, column=2).value
    assert str(last_cat) == "Grand Total"
    got = Decimal(str(last_val)).quantize(Decimal("0.01"))
    assert got == Decimal("1381.32")
